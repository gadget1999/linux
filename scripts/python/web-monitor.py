#!/usr/bin/env python3

import os, sys
import json
import time, datetime
import dateutil.parser as parser
# for struct-like class
import copy
import dataclasses
# for web APIs
import socket, ipaddress
import requests
import urllib.parse
# for reporting
import jinja2 # HTML report
import openpyxl # Excel operations
import openpyxl.styles # Excel formatting
import io
import influxdb # InfluxDB history
# for email
import email_util
# for SSL rating
import ssl_rating
# for web utilities
import web_util
# for Azure DNS glitch (use custom DNS to confirm)
import dns.resolver
dns.resolver.default_resolver = dns.resolver.Resolver(configure=False)
dns.resolver.default_resolver.nameservers = ["9.9.9.9"]
# for logging and CLI arguments parsing
import configparser
import common
import threading
logger = common.Logger.getLogger()
common.Logger.disable_http_tracing()

# some constants to handle special error conditions
POSSIBLE_DNS_GLITCH = "Name or service not known"
# add header for identity
APP_ID = "3ec1184c-03cd-4d44-b82a-0c6b14982201"

# Some times requests or socket get 'Name or service not known' incorrectly, can use a different DNS server to confirm

@dataclasses.dataclass
class SiteRecord:
  url: str
  alive: bool = False
  online: bool = False
  response_time: int = 0
  ip: str = ''
  error: str = ''
  ssl_expires: str = ''
  ssl_rating: str = ''
  ssl_report: str = ''

class SiteInfo:
  def is_valid_url(url):
    url = url.lower()
    if url.startswith('https://'):
      return True
    elif url.startswith('http://'):
      return True
    else:
      return False

  def get_status(url, allow_retry=True):
    status = SiteRecord(url=url)
    # return alive (if reachable), online (if functional) and error if any
    # by default alive and online status will be False unless explicitly set to True
    try:
      #logger.debug(f"Checking [{url}] status...")
      time.sleep(1)
      t_start = time.perf_counter_ns()
      headers = {
        "Accept-Language": "en-US,en;q=0.5",
        "User-Agent": web_util.get_user_agent(),
        "App-Id": APP_ID
      }
      r = requests.get(url, headers=headers, timeout=120)
      r.close()
      t_stop = time.perf_counter_ns()
      t_elapsed_ms = int((t_stop - t_start) / 1000000)
      status.response_time = t_elapsed_ms
      if (r.status_code < 400) or (r.status_code == 401):
        logger.debug(f"Online (status={r.status_code}, time={t_elapsed_ms}ms)")
        if (t_elapsed_ms > 10000):
          logger.error(f"{url} response time too long: {t_elapsed_ms}ms")
        status.alive = True
        status.online = True
      elif "maintenance" in r.text:
        status.alive = True
        status.online = True
        logger.info(f"{url} is under maintenance: {r.text}")
      else:
        status.error = f"HTTP error code: {r.status_code}"
        logger.error(f"{url} failed: {status.error}")
        status.alive = True
      return status
    except Exception as e:
      error_type = type(e).__name__
      error_msg = f"{e}"
      if (POSSIBLE_DNS_GLITCH in error_msg):
        if allow_retry:
          # retry once for DNS error
          time.sleep(15)
          return SiteInfo.get_status(url, False)
        logger.error(f"{url} DNS error: {POSSIBLE_DNS_GLITCH}")
        # retry still failed, try to ping IP directly (this may not be accurate for sites using reverse proxy)
        if web_util.is_host_reachable(url):
          # ignore once since requests DNS is at fault, but hard to let it use alternative DNS
          status.alive = True
          status.online = True
          return status
      else:
        logger.error(f"{url} failed: {error_type} - {error_msg}")
      status.error = f"{error_type}: {error_msg}"
      if error_type not in ['ConnectionError', 'Timeout', 'SSLError']:
        status.alive = True
      return status

  def is_blocked(url):
    try:
      time.sleep(1)
      headers = {
        "Accept-Language": "en-US,en;q=0.5",
        "User-Agent": web_util.get_user_agent()
      }
      r = requests.get(url, headers=headers)
      r.close()
      if r.status_code < 400:
        logger.error(f"Online (status={r.status_code}) --> Unexpected!")
        return False
      else:
        logger.debug(f"HTTP error code: {r.status_code} --> Expected")
        return True
    except Exception as e:
      logger.debug(f"Network error: {e} --> Expected")
      return True

  def get_report(url, include_ssl_rating=False):
    url = url.strip(' \r\'\"\n').lower()
    site_info = SiteInfo.get_status(url)
    if not site_info.alive \
       or url.startswith('http://') \
       or not include_ssl_rating:
      # no point to continue if not alive, or it's HTTP, or no need for SSL info
      return [site_info]
    # basic SSL info
    ssl_expiration_info = ssl_rating.SSLReport.get_ssl_expires_in_days(url, get_ip_addresses_func=web_util.get_ip_addresses, is_host_reachable_func=web_util.is_host_reachable)[0]
    site_info.ssl_expires = ssl_expiration_info.expires
    if ssl_expiration_info.error:
      site_info.error = ssl_expiration_info.error
    if not ssl_rating.SSLReport.should_get_rating():
      return [site_info]
    # get full SSL report
    final_reports = []
    ssl_rating_info = ssl_rating.SSLReport.get_site_rating(url)
    for record in ssl_rating_info:
      report = copy.copy(site_info)
      report.ip = record.ip
      ssl_expiration_info = ssl_rating.SSLReport.get_ssl_expires_in_days(url, record.ip, get_ip_addresses_func=web_util.get_ip_addresses, is_host_reachable_func=web_util.is_host_reachable)[0]
      report.ssl_expires = ssl_expiration_info.expires
      if ssl_expiration_info.error:
        report.error = ssl_expiration_info.error
      report.ssl_rating = record.grade
      if (not report.online) and (record.grade in ['A+', 'A', 'B']):
        # if SSL scanner can analyze, assume it's OK then
        report.online = True
      if record.error:
        # SSL rating error has higher priority
        report.error = record.error
      if record.report:
        report.ssl_report = record.report
      final_reports.append(report)
    return final_reports

@dataclasses.dataclass
class WebHookConfig:
  endpoint: str = None
  content_formatter: str = None

@dataclasses.dataclass
class InfluxDBConfig:
  endpoint: str = None
  token: str = None
  tenant: str = None
  bucket: str = None

@dataclasses.dataclass
class EmailConfig:
  email_provider: email_util.EmailProviderBase = None
  sender: str = ""
  recipients: str = ""
  subject_formatter: str = ""
  body_template: str = ""
  include_attachment: bool = False

# to start simple, this utility just do one-pass checking (no internal scheduler)
class WebMonitor:
  #########################################
  # Internal helper functions
  #########################################

  def is_future_time(self, time):
    try:
      if type(time) is datetime.datetime:
        test_time = time
      else:
        test_time = parser.parse(time)
      return True if (test_time > datetime.datetime.now()) else False
    except Exception as e:
      logger.error(f"Failed to parse maintenance time [{time}]: {e}")
      return False

  def _load_urls_from_xlsx(self, filepath):
    try:
      urls_by_sheet = {}
      workbook = openpyxl.load_workbook(filepath)
      for sheet in workbook.worksheets:
        url_count = 0
        urls_in_sheet = []
        for row in sheet.rows:
          line = row[0].value
          if not line:
            break
          line = line.lower().strip(' \r\'\"\n')
          if line.startswith(("http://", "https://")):
            # also check if there is a maintenance
            ignore_until = row[1].value
            if ignore_until and self.is_future_time(ignore_until):
              logger.debug(f"{line} is under maintenance until {ignore_until}")
              continue
            if self._include_SSL_grade and len(row) > 2:
              # check if SSL report is required for the URL
              include_ssl_grade = row[2].value
              if include_ssl_grade is None or ("yes" not in include_ssl_grade.lower()):
                continue
            urls_in_sheet.append(line)
            url_count += 1
        # Excel only logic: if there are URLs in 'Internal' tab, record them separately
        if sheet.title == 'Internal':
          logger.debug(f"Found {url_count} INTERNAL URLs")
          self._URLS_BLOCKED = urls_in_sheet
        else:
          logger.debug(f"Sheet [{sheet.title}]: found {url_count} URLs")
          urls_by_sheet[sheet.title] = urls_in_sheet
      workbook.close()
      return urls_by_sheet
    except Exception as e:
      logger.critical(f"Cannot load site list file [{filepath}]: {e}")
      sys.exit(1)

  def _load_urls_from_txt(self, filepath):
    try:
      urls = []
      with open(filepath, 'r') as f:
        lines = f.readlines()
        for line in lines:
          line = line.strip(' \r\'\"\n')
          if not line:
            continue
          if line not in urls:
            urls.append(line)
      return urls
    except Exception as e:
      logger.critical(f"Cannot load site list file [{filepath}]: {e}")
      sys.exit(1)

  def _load_urls(self, filepath):
    if filepath.lower().endswith('.xlsx'):
      return self._load_urls_from_xlsx(filepath)
    else:
      # For txt, treat as a single "Default" sheet
      return {"Default": self._load_urls_from_txt(filepath)}

  def _load_email_config(self, config, section):
    try:
      if section not in config:
        return None
      sectionConfig = config[section]
      settings = EmailConfig()
      provider = sectionConfig.get("EmailProvider", "sendgrid").strip('" ')
      api_key = os.environ["EMAIL_API_KEY"]
      settings.email_provider = email_util.get_email_provider(provider, api_key)
      settings.sender = sectionConfig.get("Sender", "").strip('" ')
      settings.recipients = sectionConfig.get("Recipients", "").strip('" ')
      settings.subject_formatter = sectionConfig.get("Subject", "").strip('" ')
      settings.include_attachment = sectionConfig.getboolean('Attachment', fallback=False)
      # use full path for body template
      template_file = sectionConfig.get("BodyTemplate", "").strip('" ')
      if template_file and template_file == os.path.basename(template_file):
        template_file = os.path.join(self._config_dir, template_file)
      settings.body_template = template_file
      return settings
    except Exception as e:
      logger.error(f"Email configuration is invalid: {e}")
      raise

  def _load_webhook_config(self, config, section):
    try:
      if section not in config:
        return None
      sectionConfig = config[section]
      settings = WebHookConfig()
      settings.endpoint = sectionConfig["EndPoint"].strip('" ')
      settings.content_formatter = sectionConfig["Content"].strip('" ')
      return settings
    except Exception as e:
      logger.error(f"WebHook configuration is invalid: {e}")
      raise

  def _load_influxdb_config(self, config, section):
    try:
      if section not in config:
        return None
      sectionConfig = config[section]
      settings = InfluxDBConfig()
      settings.endpoint = sectionConfig["InfluxDBAPIEndPoint"].strip('" ')
      settings.token = sectionConfig["InfluxDBAPIToken"].strip('" ')
      settings.tenant = sectionConfig["InfluxDBTenant"].strip('" ')
      settings.bucket = sectionConfig["InfluxDBBucket"].strip('" ')
      return settings
    except Exception as e:
      logger.error(f"InfluxDB configuration is invalid: {e}")
      raise

  def _load_sslscanner_config(self, sslscannerconfig):
    try:
      # Convert ConfigParser section to dictionary format
      config_dict = {
        "generate_rating": sslscannerconfig.getboolean("GenerateSSLRating", fallback=False),
        "use_ssllabs": sslscannerconfig.getboolean("UseSSLLabs", fallback=False)
      }
    
      if not config_dict["use_ssllabs"]:
        config_dict["local_scanner"] = sslscannerconfig.get("LocalScanner", "").strip('\" ')
        config_dict["openssl_path"] = sslscannerconfig.get("OpenSSLPath", "").strip('\" ')
        config_dict["show_progress"] = sslscannerconfig.getboolean("ShowProgress", fallback=False)

      return ssl_rating.create_ssl_config(config_dict)
    except Exception as e:
      logger.error(f"SSL scanner configuration is invalid: {e}")
      raise

  def _get_report(self, urls, include_ssl_rating=False, sheet_name=None):
    # unchanged, but now only processes a list of URLs
    full_report = []
    has_down_sites = False
    total = len(urls)
    i = 1
    for url in urls:
      if '://' not in url:
        # assume https
        url = f"https://{url}"
      if not SiteInfo.is_valid_url(url):
        logger.warning(f"Skipping invalid URL: {url}")
        continue
      tab_info = f"[{sheet_name}] " if sheet_name else ""
      logger.debug(f"Analyzing site {tab_info}({i}/{total}): {url}")
      result = SiteInfo.get_report(url, include_ssl_rating)
      i += 1
      if not result[0].online:
        has_down_sites = True
      for record in result:
        full_report.append(record)
    return full_report, has_down_sites

  def _get_report_multithreaded(self, urls_by_sheet, include_ssl_rating=False):
    threads = []
    results = {}
    errors = {}

    def worker(sheet, urls):
      try:
        report, has_down = self._get_report(urls, include_ssl_rating, sheet_name=sheet)
        results[sheet] = report
        errors[sheet] = has_down
      except Exception as e:
        logger.error(f"Thread for sheet {sheet} failed: {e}")
        results[sheet] = []
        errors[sheet] = False

    for sheet, urls in urls_by_sheet.items():
      t = threading.Thread(target=worker, args=(sheet, urls))
      t.start()
      threads.append(t)

    for t in threads:
      t.join()

    # Combine all reports and error flags
    full_report = []
    has_down_sites = False
    for sheet in urls_by_sheet:
      full_report.extend(results.get(sheet, []))
      if errors.get(sheet, False):
        has_down_sites = True
    return full_report, has_down_sites

  def _reconfirm_sites(self, report):
    has_down_sites = False
    for record in report:
      if not record.online:
        status = SiteInfo.get_status(record.url)
        record.alive = status.alive
        record.online = status.online
        record.error = status.error
        if not status.online:
          has_down_sites = True
        else:
          logger.info(f"Site is now online: {record.url}")
    return has_down_sites

  def _get_report_blocked(self, urls):
    report_blocked = []
    total = len(urls)
    i = 1
    for url in urls:
      if '://' not in url:
        # assume https
        url = f"https://{url}"
      if not SiteInfo.is_valid_url(url):
        logger.warning(f"Skipping invalid URL: {url}")
        continue
      logger.debug(f"Analyzing INTERNAL site ({i}/{total}): {url}")
      i += 1
      if SiteInfo.is_blocked(url):
        continue
      # if online, means mis-configuration
      record = SiteRecord(url=url, alive=True, online=True, error="Internal URL not blocked.")
      report_blocked.append(record)
    return report_blocked

  def _render_template(self, template, report, outputfile=None):
    """Render output based on list of SiteRecord objects"""
    engine = jinja2.Template(template)
    html = engine.render(sites=report)
    if outputfile:
      with open(outputfile, 'w') as f:
        f.write(html)
    return html

  def _generate_xlsx_report(self, report, outputfile=None):
    """Generate Excel report with site report data using openpyxl"""
    # Create a new workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Site Report'
    
    # Create header with formatting
    header_font = openpyxl.styles.Font(bold=True)
    good_font = openpyxl.styles.Font(bold=True, color="00800000")  # Green
    bad_font = openpyxl.styles.Font(bold=True, color="00FF0000")   # Red
    
    headers = ['On', 'Grade', 'Expires In (days)', 'URL', 'IP', 'Error', 'City', 'Region', 'Country']
    column_widths = [4, 6, 18, 40, 15, 40, 10, 10, 8]
    
    # Set headers
    for col, (header, width) in enumerate(zip(headers, column_widths), 1):
      cell = worksheet.cell(row=1, column=col, value=header)
      cell.font = header_font
      worksheet.column_dimensions[cell.column_letter].width = width
    
    # Add auto-filter
    worksheet.auto_filter.ref = f"A1:{worksheet.cell(row=1, column=len(headers)).coordinate}"
    
    # Fill in sheet with report data
    for row_idx, record in enumerate(report, 2):  # Start from row 2
      # Online status
      online_cell = worksheet.cell(row=row_idx, column=1, value='Y' if record.online else 'N')
      online_cell.font = good_font if record.online else bad_font
      
      # SSL Grade
      if record.ssl_rating:
        grade_cell = worksheet.cell(row=row_idx, column=2, value=record.ssl_rating)
        grade_cell.font = good_font if record.ssl_rating.startswith('A') else bad_font
      
      # SSL Expires
      if record.ssl_expires:
        expires_cell = worksheet.cell(row=row_idx, column=3, value=record.ssl_expires)
        expires_cell.font = good_font if record.ssl_expires > 60 else bad_font
      
      # URL (with hyperlink if SSL report available)
      url_cell = worksheet.cell(row=row_idx, column=4, value=record.url)
      if record.ssl_report:
        url_cell.hyperlink = record.ssl_report
        url_cell.style = "Hyperlink"
      
      # IP Address
      worksheet.cell(row=row_idx, column=5, value=record.ip)
      
      # Error
      if record.error:
        error_cell = worksheet.cell(row=row_idx, column=6, value=record.error)
        error_cell.font = bad_font
      
      # Location columns (currently not populated as noted in original code)
      # Columns 7-9 for City, Region, Country left empty as in original
    
    # Save or return as bytes
    if outputfile:
      workbook.save(outputfile)
      workbook.close()
      return None
    else:
      # Save to BytesIO for in-memory handling
      with io.BytesIO() as output:
        workbook.save(output)
        workbook.close()
        output.seek(0)
        return output.read()

  def _send_email_report(self, report):
    try:
      if not self._email_settings:
        logger.warning("Email settings not configured, skipping email report.")
        return

      # Format subject with current date/time
      today = time.strftime('%Y-%m-%d', time.localtime())
      now = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
      mapping = {'now': now, 'today': today}
      subject = self._email_settings.subject_formatter.format_map(mapping)

      # Load email body template
      with open(self._email_settings.body_template, 'r') as f:
        template_content = f.read()
      
      # Render HTML content
      html_content = self._render_template(template_content, report)
      
      # Generate Excel attachment if enabled
      attachment_data = None
      attachment_filename = None
      attachment_type = None
      
      if self._email_settings.include_attachment:
        attachment_data = self._generate_xlsx_report(report)
        attachment_filename = f"{today}-Site-Report.xlsx"
        attachment_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
      success = self._email_settings.email_provider.send_email(
        sender=self._email_settings.sender,
        recipients=self._email_settings.recipients,
        subject=subject,
        html_content=html_content,
        attachment_data=attachment_data,
        attachment_filename=attachment_filename,
        attachment_type=attachment_type
      )
      if not success:
        logger.error("Failed to send email report")
    except Exception as e:
      logger.error(f"Failed to send Site Report: {e}")

  def _send_webhook_notice(self, report):
    # get post body
    content = "Following sites may be down:<br>"
    for record in report:
      if not record.online:
        content += f"{record.url} ({record.error})<br>"

    # contruct payload
    webhook_config = self._webhook_settings
    # make string json safe
    content = content.replace('\r', '').replace('\n', '').replace('\\', '\\\\').replace('"', '\\"')
    payload = webhook_config.content_formatter.format(content=content)

    # send over notification
    try:
      logger.debug("Sending webhook notification...")
      headers = {
        "Content-Type": "application/json",
        "User-Agent": web_util.get_user_agent()
      }
      r = requests.post(webhook_config.endpoint, headers=headers, data=payload)
      if r.status_code > 400:
        logger.error(f"Post to webhook failed: {r.status_code}")
    except Exception as e:
      logger.error(f"Post to webhook failed: {e}")

  def _store_influxdb_report(self, report):
    influxdb_settings = InfluxDBConfig(
      endpoint=self._influxdb_settings.endpoint,
      token=self._influxdb_settings.token,
      tenant=self._influxdb_settings.tenant,
      bucket=self._influxdb_settings.bucket
      )
    influxdb_writer = influxdb.InfluxDBHelper(influxdb_settings)
    logger.debug("Storing metrics into InfluxDB...")
    for record in report:
      parsed_uri = urllib.parse.urlparse(record.url)
      data = []
      if not record.error:
        data.append(("Response_Time", record.response_time))
      data.append(("Offline", 0 if record.online else 1))
      try:
        influxdb_writer.report_data_list("Metrics", parsed_uri.hostname, data)
      except Exception as e:
        logger.error(f"Failed to store InfluxDB record: {parsed_uri.hostname}: {e}")


  #########################################
  # Public methods
  #########################################

  def __init__(self, configfile):
    try:
      if not os.path.isfile(configfile):
        raise Exception(f"Config file [{configfile}] does not exist.")
      config = configparser.ConfigParser()
      config.read(configfile)
      self._config_dir = os.path.dirname(configfile)
      self._retry_delay = config.getint("Global", "RetryDelay", fallback=120)
      self._max_retries = config.getint("Global", "MaxRetries", fallback=5)
      self._include_SSL_report = config.getboolean("SSL", "GetSSLReport", fallback=False)
      self._include_SSL_grade = config.getboolean("SSL", "GenerateSSLRating", fallback=False)
      url_list_file = config["Global"]["URLFile"]
      if url_list_file == os.path.basename(url_list_file):
        url_list_file = os.path.join(self._config_dir, url_list_file)
      self._URLs_by_sheet = self._load_urls(url_list_file)
      self._email_settings = self._load_email_config(config, "Email")
      self._influxdb_settings = self._load_influxdb_config(config, "InfluxDB")
      self._webhook_settings = self._load_webhook_config(config, "WebHook")
      if self._include_SSL_report:
        ssl_rating.SSLReport.set_config(self._load_sslscanner_config(config["SSL"]))
    except Exception as e:
      logger.error(f"Config file {configfile} is invalid: {e}")
      raise

  def check_sites(self):
    full_report, has_down_sites = self._get_report_multithreaded(self._URLs_by_sheet, self._include_SSL_report)
    # reconfirm failed sites
    retries = 0
    while has_down_sites and retries < self._max_retries:
      retries += 1
      logger.info(f"Wait some time and retry (#{retries}) failed sites.")
      time.sleep(self._retry_delay)
      has_down_sites = self._reconfirm_sites(full_report)
    if len(full_report) == 0:
      logger.error(f"Site report list is empty.")
      return
    # check if any of blocked sites are accessible
    if '_URLS_BLOCKED' in dir(self):
      full_report.extend(self._get_report_blocked(self._URLS_BLOCKED))
    # sort list to move items with error to front
    full_report.sort(key=lambda i: i.error if i.error else '', reverse=True)
    full_report.sort(key=lambda i: i.ssl_rating if i.ssl_rating else 'Unknown', reverse=True)
    full_report.sort(key=lambda i: i.online)
    full_report.sort(key=lambda i: i.ssl_expires if i.ssl_expires else 0)
    num_errors = sum(1 for x in full_report if x.error)
    # always record metrics stats
    if self._influxdb_settings:
      self._store_influxdb_report(full_report)
    # send email if ssl rating included, or has failed sites, or has errors
    if self._include_SSL_report or has_down_sites or num_errors > 0:
      if self._email_settings:
        self._send_email_report(full_report)
      if self._webhook_settings:
        self._send_webhook_notice(full_report)
      # also archive the report locally, in case email gets lost
      now = datetime.datetime.now()
      archive_folder = '/tmp/dropbox_archive'
      if not os.path.exists(archive_folder):
        os.makedirs(archive_folder)
      report_file = f"{archive_folder}/Site-Report-{now.strftime('%Y-%m-%d_%H_%M_%S')}.xlsx"
      
      # Generate and save Excel report
      self._generate_xlsx_report(full_report, report_file)
      if num_errors > 0:
        logger.error(f"Scan completed: {num_errors} of {len(full_report)} URLs have errors.")
      else:
        logger.info(f"Scan completed: no errors for {len(full_report)} URLs.")

########################################
# CLI interface
########################################

def check_sites(args):
  web_util.get_latest_user_agent()
  monitor = WebMonitor(args.config)
  monitor.check_sites()

#################################
# Program starts
#################################
if (__name__ == '__main__') and ('UNIT_TEST' not in os.environ):
  CLI_config = { 'func':check_sites, 'arguments': [
    {'name':'config', 'help':'Config file for monitor'} 
    ]}
  common.CLIParser.run(CLI_config)
